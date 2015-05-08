## Excel
# Don't use
# https://pythonhosted.org/openpyxl/usage.html#inserting-an-image

from openpyxl import Workbook
from openpyxl.drawing import Image


wb = Workbook()
ws = wb.active
ws['A1'] = 'You should see three logos below'
ws['A2'] = 'Resize the rows and cells to see anchor differences'

# create image instances
img = Image('logo.png')
img2 = Image('logo.png')
img3 = Image('logo.png')

# place image relative to top left corner of spreadsheet
img.drawing.top = 100
img.drawing.left = 150

# the top left offset needed to put the image
# at a specific cell can be automatically calculated
img2.anchor(ws['D12'])

# one can also position the image relative to the specified cell
# this can be advantageous if the spreadsheet is later resized
# (this might not work as expected in LibreOffice)
img3.anchor(ws['G20'], anchortype='oneCell')

# afterwards one can still add additional offsets from the cell
img3.drawing.left = 5
img3.drawing.top = 5

# add to worksheet
ws.add_image(img)
ws.add_image(img2)
ws.add_image(img3)
wb.save('logo.xlsx')